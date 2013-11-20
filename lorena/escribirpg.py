import openpyxl
#!/usr/bin/env python
# *.* coding=utf-8 *.*

#from pg8000 import DBAPI
#conn = DBAPI.connect(host="161.196.204.6", user="postgres", password="postgres",port=5433,database="prueba")
#cursor = conn.cursor()

libro = openpyxl.load_workbook(filename = 'Libros.xls',guess_types=True)






"""
hoja = libro.get_active_sheet()
for i in xrange(3,56):
    title=hoja.cell(row=i,col=1)
    author=hoja.cell(row=i,col=2)
    tip=hoja.cell(row=i,col=3)
    pagina=hoja.cell(row=i,col=4)
    print title,author,tip,pagina
"""




